from datetime import datetime, timedelta
from unidecode import unidecode  # Importa la función unidecode
from django.contrib import messages, auth
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Sum, Q, F
from django.db.models.functions import Lower
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render, redirect
from django.template.loader import get_template
from django.urls import reverse, reverse_lazy
from django.utils.decorators import method_decorator
from django.views import View
from django.views.generic import UpdateView, ListView, TemplateView, View
from .choices import ANIO_CHOICES, MES_CHOICES

# Generar pdf
from openpyxl import Workbook
from xhtml2pdf import pisa

from .forms import FormNuevoReporte, UserForm, UserRegisterForm, ReciboSelectContratacionForm
from .models import Cliente, DetallePago, Informacion, Servicio, Contratacion, Recibo, ReporteFallo


def template_view(request):
    return render(request, 'base_template/index.html')


# Función de pertenencia a grupos individuales o en colección
def usuario_pertenece_grupos(id_usuario, nombres_grupos):
    try:
        usuario = User.objects.get(pk=id_usuario)
        grupos_usuario = usuario.groups.values_list('name', flat=True)
        pertenece = any(nombre_grupo in grupos_usuario for nombre_grupo in nombres_grupos)
    except User.DoesNotExist:
        pertenece = False
    return pertenece


# Vistas para inicio y cierre de sesión
def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = auth.authenticate(username=username, password=password)
        if user is not None:
            auth.login(request, user)
            messages.success(request, 'Inicio de sesión correcto')
            return redirect(
                'home:home')  # Cambia 'home' por la URL de la página a la que deseas redirigir después del inicio de sesión exitoso
        else:
            messages.error(request, 'Credenciales inválidas')

    return render(request, 'registration/login.html')


@login_required
def HomeView(request):
    enddate = datetime.today() + timedelta(days=1)
    startdate = enddate - timedelta(days=7)
    recibos = Recibo.objects.filter(fecha__range=[startdate, enddate])
    
    contrataciones = Contratacion.objects.all()
    deudores = contrataciones.filter(estado='P')
    deudores_porcentaje =  (deudores.count()/contrataciones.count())*100
    
    reportes = ReporteFallo.objects.filter(~Q(estado='S'))
    
    ingresoSemana = recibos.aggregate(Sum('total'))['total__sum']
    if ingresoSemana is None:
        ingresoSemana = 0
    return render(request, 'index.html',
                  {'contrataciones': contrataciones, 'recibos': recibos, 'ingresototal': ingresoSemana,
                   'deudores': deudores_porcentaje, 'reportes':reportes})


def LogoutView(request):
    logout(request)
    return redirect('home:login')


# ----------------- Vistas CRUD Clientes -----------------
# Vista Para Registrar Clientes
@login_required
def NuevoClienteView(request):
    cui = request.POST['cui']
    nombre = request.POST['nombre']
    apellido = request.POST['apellido']
    telefono = request.POST['telefono']
    correo = request.POST['correo']

    cliente = Cliente(cui=cui, nombre=nombre, apellido=apellido, telefono=telefono, correo=correo)
    cliente.save()

    return redirect('home:clientes')


# Vista Para Listar Clientes
@method_decorator(login_required, name='dispatch')
class ClienteListView(ListView):
    model = Cliente
    template_name = 'Clientes.html'
    context_object_name = 'clientes'

    def get_queryset(self):
        query = self.request.GET.get('nombre')
        if query:
            queryset = Cliente.objects.filter(nombre__icontains=query) | Cliente.objects.filter(apellido__icontains=query)
        else:
            queryset = Cliente.objects.all()
        return queryset



# Vista Para Borrar Clientes
def BorrarClienteView(request, pk):
    cliente = Cliente.objects.get(pk=pk)
    cliente.delete()
    return redirect('home:clientes')


# Vista Para Modificar Clientes
class ModificarClienteView(UpdateView):
    template_name = 'modificarCliente.html'
    model = Cliente
    fields = ['cui', 'nombre', 'apellido', 'direccion', 'telefono', 'correo']
    success_url = reverse_lazy('home:clientes')


# Vistas para listar servicios
def ServiciosView(request):
    return render(request, 'servicios.html', {'servicios': Servicio.objects.all()})


# Vista para insertar servicios
def NuevoServicioView(request):
    tipo = request.POST['tipo']
    nombre = request.POST['nombre']
    ancho_banda = request.POST['ancho_banda']
    costo = request.POST['costo']
    servicio = Servicio(tipo=tipo, nombre=nombre, ancho_banda=ancho_banda, costo=costo)
    servicio.save()
    return redirect('home:servicios')


# Vista para borrar servicios
def BorrarServicioView(request, pk):
    servicio = Servicio.objects.get(pk=pk)
    servicio.delete()
    return redirect('home:servicios')


# Vista para modificar servicios
class ModificarServicioView(UpdateView):
    template_name = 'modificarServicio.html'
    model = Servicio
    fields = ['tipo', 'nombre', 'ancho_banda', 'costo']
    success_url = reverse_lazy('home:servicios')


# Vista para listar contrataciones
def ContratacionesView(request):
    return render(request, 'contrataciones.html', {
        'contrataciones': Contratacion.objects.all(),
        'clientes': Cliente.objects.all(),
        'servicios': Servicio.objects.all()
    })


# Vista para crear una contratacion
def NuevaContratacionView(request):
    cliente = Cliente.objects.get(pk=(request.POST['cliente']))
    servicio = Servicio.objects.get(pk=(request.POST['servicio']))
    direccion = request.POST['direccion']
    contratacion = Contratacion(cliente=cliente, servicio=servicio, direccion=direccion)
    contratacion.save()
    return redirect('home:contrataciones')


# Vista para borrar contratacion
def BorrarContratacionView(request, pk):
    contratacion = Contratacion.objects.get(pk=pk)
    contratacion.delete()
    return redirect('home:contrataciones')


# Vista para modificar contratacion
class ModificarContratacionView(UpdateView):
    template_name = 'modificarContratacion.html'
    model = Contratacion
    fields = ['cliente', 'servicio', 'direccion']
    success_url = reverse_lazy('home:contrataciones')


class ContratacionAPIView(View):
    def get(self, request, *args, **kwargs):
        # Obtén el valor del parámetro 'q' de la solicitud
        
        search_term = request.GET.get('q')

        # Convierte el término de búsqueda a minúsculas y elimina tildes
        search_term = unidecode(search_term.lower())
        print(search_term)
        
        # Realiza una consulta para obtener las contrataciones cuyo cliente tiene un nombre o apellido que contiene el término de búsqueda
        contrataciones = Contratacion.objects.filter(
            Q(cliente__nombre__icontains=search_term) | Q(cliente__apellido__icontains=search_term)
        )

        # Convierte las contrataciones en un formato JSON
        contrataciones_json = [{'id': c.id, 'cliente': str(c.cliente), 'servicio': str(c.servicio),}
                              for c in contrataciones]

        # Devuelve la respuesta JSON
        return JsonResponse({'contrataciones': contrataciones_json})

    def post(self, request, *args, **kwargs):
        # Aquí puedes manejar la lógica para crear una nueva contratación
        # Recupera los datos del cuerpo de la solicitud (request body) y crea una nueva Contratacion
        # Por ejemplo, si los datos se envían como JSON en el cuerpo de la solicitud:
        data = request.POST  # Asegúrate de que los datos se envíen como JSON
        cliente_id = data.get('cliente')
        servicio_id = data.get('servicio')
        direccion = data.get('direccion')
        saldo = data.get('saldo')
        ultimo_pago = data.get('ultimo_pago')
        estado = data.get('estado')

        # Crea una nueva Contratacion con los datos proporcionados
        nueva_contratacion = Contratacion(
            cliente_id=cliente_id,
            servicio_id=servicio_id,
            direccion=direccion,
            saldo=saldo,
            ultimo_pago=ultimo_pago,
            estado=estado
        )
        nueva_contratacion.save()

        # Después de crear la nueva contratación, puedes devolver una respuesta JSON
        return JsonResponse({'message': 'Contratación creada con éxito'})

    def put(self, request, pk, *args, **kwargs):
        # Aquí puedes manejar la lógica para actualizar una contratación existente
        # pk es el ID de la contratación que deseas actualizar
        # Recupera los datos del cuerpo de la solicitud (request body) y actualiza la Contratacion
        # Por ejemplo, si los datos se envían como JSON en el cuerpo de la solicitud:
        data = request.POST  # Asegúrate de que los datos se envíen como JSON
        cliente_id = data.get('cliente')
        servicio_id = data.get('servicio')
        direccion = data.get('direccion')
        saldo = data.get('saldo')
        ultimo_pago = data.get('ultimo_pago')
        estado = data.get('estado')

        # Obtiene la Contratacion existente por su ID
        contratacion = Contratacion.objects.get(pk=pk)

        # Actualiza los campos de la Contratacion con los nuevos datos
        contratacion.cliente_id = cliente_id
        contratacion.servicio_id = servicio_id
        contratacion.direccion = direccion
        contratacion.saldo = saldo
        contratacion.ultimo_pago = ultimo_pago
        contratacion.estado = estado

        # Guarda los cambios en la Contratacion
        contratacion.save()

        # Después de actualizar la contratación, puedes devolver una respuesta JSON
        return JsonResponse({'message': 'Contratación actualizada con éxito'})

    def delete(self, request, pk, *args, **kwargs):
        # Aquí puedes manejar la lógica para eliminar una contratación
        # pk es el ID de la contratación que deseas eliminar
        # Realiza la eliminación de la Contratacion
        # Por ejemplo:
        try:
            contratacion = Contratacion.objects.get(pk=pk)
            contratacion.delete()
            return JsonResponse({'message': 'Contratación eliminada con éxito'})
        except Contratacion.DoesNotExist:
            return JsonResponse({'message': 'La Contratación no existe'}, status=404)


# Vista para registrar informacion de empresa
def InformacionView(request):
    nombre = request.POST['nombre']
    direccion = request.POST['direccion']
    telefono = request.POST['telefono']
    informacion = Informacion.objects.get(pk=1)
    informacion.nombre = nombre
    informacion.direccion = direccion
    informacion.telefono = telefono
    informacion.save()

    return redirect('home:informacion')


# Vista para listar informacion de empresa
def informacionempresa_view(request):
    form = UserRegisterForm()
    context = {'informacion': Informacion.objects.get(pk=1), 'form': form,
               'usuarios': User.objects.all().order_by('first_name')}
    return render(request, 'informacion.html', context)


# ----------------- Vistas CRUD Usuarios -----------------
def usuarios_view(request):
    form = UserRegisterForm()
    context = {'form': form, 'usuarios': User.objects.all()}
    return render(request, 'usuarios/usuarios.html', context)


def usuarios_filtrados(request):
    form = UserRegisterForm()
    nombre = request.GET.get('nombre', '')  # Obtener el valor del parámetro 'nombre' de la solicitud GET

    usuarios = User.objects.filter(Q(first_name__icontains=nombre) | Q(
        last_name__icontains=nombre))  # Filtrar usuarios cuyos nombres contengan el valor proporcionado

    context = {'informacion': Informacion.objects.get(pk=1), 'form': form, 'usuarios': usuarios}

    return render(request, 'informacion.html', context)


def usuarionuevo_view(request):
    form = UserRegisterForm(request.POST)
    if form.is_valid():
        form.save()
        username = form.cleaned_data['first_name']
        messages.success(request, f'Usuario {username} creado correctamente')
    else:
        error_message = form.errors.as_text()
        messages.error(request, f'El usuario no se creó. Error: {error_message}')
    return redirect('home:usuarios')


def usuarioeliminar_view(request, username):
    colaborador = User.objects.get(username=username)
    colaborador.delete()
    messages.success(request, f'Usuario {colaborador.first_name} eliminado correctamente')
    return redirect('home:usuarios')


class EditarUsuarioView(UpdateView):
    template_name = 'usuarios/usuario_editar.html'
    form_class = UserForm
    success_url = reverse_lazy('home:usuarios')
    model = User


# Vista para borrar informacion de empresa
def BorrarInformacionView(request, pk):
    informacion = Informacion.objects.get(pk=pk)
    informacion.delete()
    return redirect('home:informacion')


# Vista para modificar informacion de empresa
class ModificarInformacionView(UpdateView):
    template_name = 'modificarInformacion.html'
    model = Informacion
    fields = ['nombre', 'direccion', 'telefono']
    success_url = reverse_lazy('home:informacion')


# ----------------- Vistas CRUD Reportes de fallos -----------------
# Listado de Reportes de fallos
@login_required
def reporte_fallo(request):
    reportes = ReporteFallo.objects.all().order_by('-fecha_reporte')
    usuarios = User.objects.filter(groups__name__in=['Tecnico', 'Administrador'])

    if request.method == 'POST':
        form = FormNuevoReporte(request.POST)
        if form.is_valid():
            reporte = form.save(commit=False)
            reporte.usuario = request.user
            reporte.save()
            messages.success(request, 'Se ha creado el reporte de fallo correctamente.')
            return redirect('home:reportes')  # Redirige a la vista deseada después de crear el reporte
        else:
            messages.error(request, 'Ha ocurrido un error al crear el reporte de fallo.')
    else:
        form = FormNuevoReporte

    return render(request, 'reporte_fallo.html', {'reportes': reportes, 'usuarios': usuarios, 'form': form})


# Cambio de estado de Reporte de fallo
@login_required
def cambiar_estado_reporte_fallo_view(request, pk, estado):
    if usuario_pertenece_grupos(request.user.id, ['Administrador', 'Tecnico']):
        tecnico = User.objects.get(pk=request.user.id)
        reporte = ReporteFallo.objects.get(pk=pk)
        if reporte.estado == 'S':
            messages.error(request, f'{reporte} - Ya está completado')
        else:
            reporte.estado = estado
            reporte.tecnico_asignado = tecnico
            reporte.save()
            messages.success(request, f'{reporte} - {reporte.get_estado_display()}')
    else:
        messages.error(request, 'Usuario sin autorización para cambios de estado')
    return redirect(reverse('home:reportes'))


# Cambio de técnico de Reporte de fallo
@login_required
def cambiar_tecnico_reporte_fallo_view(request, id_reporte, id_tecnico):
    if usuario_pertenece_grupos(request.user.id, ['Administrador']):
        reporte = ReporteFallo.objects.get(pk=id_reporte)
        if reporte.estado == 'S':
            messages.warning(request, f'{reporte} ya fue completado, no se puede cambiar el técnico asignado')
        else:
            tecnico = User.objects.get(pk=id_tecnico)
            reporte.tecnico_asignado = tecnico
            reporte.save()
            messages.success(request, f'{reporte} asignado a {tecnico}')
    else:
        messages.error(request, 'Usuario sin autorización para cambio de técnico')
    return redirect(reverse('home:reportes'))


# Borrar reporte de fallos
@login_required
def borrar_reporte_fallos_view(request, id_reporte):
    if usuario_pertenece_grupos(request.user.id, ['Administrador']):
        reporte = ReporteFallo.objects.get(pk=id_reporte)
        reporte.delete()
        messages.success(request, f'{reporte} eliminado')
    else:
        messages.error(request, 'Usuario sin autorización para cambio de técnico')
    return redirect(reverse('home:reportes'))


# Vista para imprimir PDF de factura
class ReciboPDFView(View):
    def get(self, request, *args, **kwargs):
        try:
            template = get_template('recibopdf.html')
            context = {
                'recibo': Recibo.objects.get(pk=self.kwargs['pk'])
            }
            html = template.render(context)
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="report.pdf"'
            # create a pdf
            pisa_status = pisa.CreatePDF(
                html, dest=response)
            return response
        except:
            pass
        return HttpResponseRedirect(reverse_lazy('home:pagos'))


# ----------------- Vistas CRUD Recibos y pagos -----------------
# Vista para listar pagos
def PagosView(request):
    form = ReciboSelectContratacionForm
    return render(request, 'pagos.html',
                  {'pagos': Recibo.objects.all().order_by('-pk'), 'contrataciones': Contratacion.objects.all(), 'form': form})
    

def NuevoRecibo(request):
    pk = request.POST['contratacion']
    contratacion = Contratacion.objects.get(pk=pk)
    recibo = Recibo(contratacion=contratacion, total=0)
    recibo.save()
    return render(request, 'recibo.html',
                  {'contratacion': contratacion, 'recibo': Recibo.objects.last(), 'meses': MES_CHOICES,
                   'anios': ANIO_CHOICES})


def NuevoDetalle(request, pk):
    recibo = pk
    mes = request.POST.get('mes', '')
    anio = request.POST.get('anio', '')
    subtotal = request.POST.get('subtotal', '')

    valido = False  # Variable que sirve para verificar si ese mes ya está pagado.

    recibo = Recibo.objects.get(pk=recibo)
    contratacion = Contratacion.objects.get(pk=recibo.contratacion.id)

    if mes != '' and anio != '' and subtotal != '':
        if not DetallePago.objects.filter(recibo__contratacion=contratacion, mes=mes, anio=anio).exists():
            detallepago = DetallePago.objects.create(mes=mes, anio=anio, subtotal=subtotal, recibo=recibo)
            contratacion.ultimo_pago += timedelta(days=30)
            contratacion.save()
            actualizar_pendiente(contratacion)

        recibo.total = DetallePago.objects.filter(recibo=recibo).aggregate(total=Sum('subtotal'))['total'] or 0
        recibo.save()
        
    return render(request, 'recibo.html', {'contratacion': contratacion, 'recibo': recibo, 'meses': MES_CHOICES,
                                           'anios': ANIO_CHOICES})

#TODO Revisar que reste cuando se elimina un detalle
@login_required
def borrar_detalle_pago_view(request, pk):
    detalle = DetallePago.objects.get(pk=pk)
    id_recibo = detalle.recibo.id # Conservar el id del recibo para volver al detalle
    detalle.delete()
    return redirect('home:nuevodetalle', pk=id_recibo)


def actualizar_pendiente(contratacion):
    dias_ultimo_pago = datetime.now().astimezone(contratacion.ultimo_pago.tzinfo) - contratacion.ultimo_pago
    # Verifica si tiene más de 30 días el último pago para sumarle a su saldo pendiente
    if dias_ultimo_pago.days >= 30:
        meses = dias_ultimo_pago / 30
        contratacion.saldo = meses.days * contratacion.servicio.costo
        contratacion.estado = 'P'  # Establece como "Pendiente de pago"
    else:
        contratacion.saldo = 0
        contratacion.estado = 'D'  # Establece como "Al día"
    contratacion.save()


class ReporteExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        cliente = Cliente.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE CLIENTES'

        ws.merge_cells('B1:G1')

        ws['B3'] = 'CUI'
        ws['C3'] = 'NOMBRE'
        ws['D3'] = 'APELLIDO'
        ws['E3'] = 'DIRECCION'
        ws['F3'] = 'TELEFONO'
        ws['G3'] = 'CORREO'

        cont = 4
        for cli in cliente:
            ws.cell(row=cont, column=2).value = cli.cui
            ws.cell(row=cont, column=3).value = cli.nombre
            ws.cell(row=cont, column=4).value = cli.apellido
            ws.cell(row=cont, column=5).value = cli.direccion
            ws.cell(row=cont, column=6).value = cli.telefono
            ws.cell(row=cont, column=7).value = cli.correo
            cont += 1

        nombre_archivo = "ReporteClientes.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ReporteContrataciones(TemplateView):
    def get(self, request, *args, **kwargs):
        contrato = Contratacion.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE CONTRATACIONES'

        ws.merge_cells('B1:E1')

        ws['B3'] = 'CLIENTE'
        ws['C3'] = 'SERVICIO'
        ws['D3'] = 'DIRECCION'
        ws['E3'] = 'CREACION'

        cont = 4
        for puesto in contrato:
            ws.cell(row=cont, column=2).value = puesto.cliente
            ws.cell(row=cont, column=3).value = puesto.servicio
            ws.cell(row=cont, column=4).value = puesto.direccion
            ws.cell(row=cont, column=5).value = puesto.creacion

            cont += 1

        nombre_archivo = "ReporteContratacion.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
